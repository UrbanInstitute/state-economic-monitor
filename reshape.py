from json import dump, load
from random import random
import time
import openpyxl
import csv
import datetime
import os
import re
import zipfile
from tempfile import NamedTemporaryFile
import shutil


DATE_FORMAT = "%Y-%m-%d"
rootPath = "/var/www/html/semapp/"
# rootPath = "/Users/silke/Repos/state-economic-monitor/"

# R's openxlsx sometimes writes a worksheet rels entry pointing to a
# drawing part that isn't actually included in the archive. Strict openpyxl
# versions raise KeyError on load. We only read cell values, so drop any
# drawing relationships before handing the file to openpyxl.
_DRAWING_REL_RE = re.compile(rb'<Relationship[^>]*Type="[^"]*/drawing"[^>]*/>')


def loadWorkbookSafely(path):
	tmp = NamedTemporaryFile(suffix=".xlsx", delete=False)
	tmp.close()
	try:
		with zipfile.ZipFile(path, "r") as zin, \
				zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zout:
			for item in zin.infolist():
				data = zin.read(item.filename)
				if item.filename.endswith(".rels"):
					data = _DRAWING_REL_RE.sub(b"", data)
				zout.writestr(item, data)
		return openpyxl.load_workbook(tmp.name, data_only=True)
	finally:
		os.remove(tmp.name)

def cleanExcelRow(row, isDate, colCount):
	if isDate:
		for i, r in enumerate(row):
			if i >= colCount:
				del row[i:len(row)]
				break
			if r == "Geography":
				continue
			else:
				if str(r).find("Q") != -1:
					year = r.split(" ")[0]
					month = "%02d" % (((int(r.split(" ")[1].replace("Q","")) - 1) * 3) + 1,)
					row[i] = year + "-" + month + "-01"
				else:
					print(r)
					print(row)
					if isinstance(r, datetime.datetime):
						pyDate = r
					else:
						pyDate = datetime.datetime.strptime(str(r), "%m/%d/%Y")
					row[i] = pyDate.strftime(DATE_FORMAT)
		return row
	else:
		for i, r in enumerate(row):
			if i >= colCount:
				del row[i:len(row)]
				break
		if row[0] == "US Average":
			row[0] = "United States"
			return row
		if row[0] == "" or row[0] is None:
			return row
		else:
			return row


def buildCSVs(indicator):
	wb = loadWorkbookSafely(rootPath + 'static/data/source/%s.xlsx' % indicator)

	if(indicator != "house_price_index"):
		raw = wb.worksheets[0]

		fileName = indicator
		if(indicator == "federal_public_employment" or indicator == "private_employment" or indicator == "public_employment" or indicator == "state_and_local_public_employment" or indicator == "total_employment" or indicator == "state_and_local_public_education_employment" or indicator == "leisure_and_hospitality_employment" or indicator == "manufacturing_employment" or indicator == "retail_trade_employment"):
			fileName += "_raw_in_thousands"
		elif(indicator == "state_gdp" or indicator == "accommodation_and_food_services_state_gdp" or indicator == "retail_trade_state_gdp" or indicator == "government_state_gdp" or indicator == "manufacturing_state_gdp" ):
			fileName += "_raw_in_millions"
		else:
			fileName += "_raw"

		with open(rootPath + 'static/data/csv/%s.csv' % fileName, 'w', newline='') as rawFile:
			rawWriter = csv.writer(rawFile, quoting=csv.QUOTE_ALL)

			# Count non-empty cells in row 6 (0-indexed row 5)
			row6 = [cell.value for cell in raw[6]]
			colCount = len(list(filter(lambda x: x is not None and x != "", row6)))

			for rownum in range(4, raw.max_row + 1):
				row = [cell.value for cell in raw[rownum]]
				if(row[0] is None or row[0] == ""):
					break
				else:
					isDate = (rownum == 4)
					rawWriter.writerow(cleanExcelRow(row, isDate, colCount))

	if(indicator != "unemployment_rate"):
		if(indicator == "house_price_index"):
			sheetIndex = 0
		else:
			sheetIndex = 1
		change = wb.worksheets[sheetIndex]

		with open(rootPath + 'static/data/csv/%s_yoy_percent_change.csv' % indicator, 'w', newline='') as changeFile:
			changeWriter = csv.writer(changeFile, quoting=csv.QUOTE_ALL)

			row6 = [cell.value for cell in change[6]]
			colCount = len(list(filter(lambda x: x is not None and x != "", row6)))

			for rownum in range(4, change.max_row + 1):
				row = [cell.value for cell in change[rownum]]
				if(row[0] is None or row[0] == ""):
					break
				else:
					isDate = (rownum == 4)
					changeWriter.writerow(cleanExcelRow(row, isDate, colCount))


with open(rootPath + "static/data/mapping/state_fips.csv", 'r') as fipsFile:
	fipsReader = csv.reader(fipsFile)
	nameToFips = {}
	for row in fipsReader:
		nameToFips[row[2]] = { "abbr": row[1], "fips": row[0], "name": row[2] }


indicators = ["federal_public_employment", "house_price_index", "private_employment", "public_employment", "state_and_local_public_employment", "state_gdp", "total_employment", "unemployment_rate", "weekly_earnings","state_and_local_public_education_employment","leisure_and_hospitality_employment","manufacturing_employment","retail_trade_employment","accommodation_and_food_services_state_gdp","retail_trade_state_gdp","government_state_gdp","manufacturing_state_gdp"]

tempDict = {}
terminalDates = {}


for indicator in indicators:
	buildCSVs(indicator)

for index, indicator in enumerate(indicators):
	if(indicator != "house_price_index"):
		key = str(index) + "r"
		fileName = indicator
		if(indicator == "federal_public_employment" or indicator == "private_employment" or indicator == "public_employment" or indicator == "state_and_local_public_employment" or indicator == "total_employment" or indicator == "state_and_local_public_education_employment" or indicator == "leisure_and_hospitality_employment" or indicator == "manufacturing_employment" or indicator == "retail_trade_employment" ):
			fileName += "_raw_in_thousands"
		elif(indicator == "state_gdp" or indicator == "accommodation_and_food_services_state_gdp" or indicator == "retail_trade_state_gdp" or indicator == "government_state_gdp" or indicator == "manufacturing_state_gdp" ):
			fileName += "_raw_in_millions"
		else:
			fileName += "_raw"

		with open(rootPath + "static/data/csv/%s.csv" % fileName, 'r') as rawCsvFile:
			rawReader = csv.reader(rawCsvFile)
			rows = list(rawReader)

		dates = rows[0]
		terminalDates[key] = {"firstDate" : dates[1], "lastDate" : dates[len(dates) - 1]}

		for row in rows[1:]:
			name = row[0]
			if(name == ""):
				print(indicator)

			for i in range(1, len(row)):
				date = dates[i]
				abbr = nameToFips[name]["abbr"]
				value = row[i]
				tempKey = abbr + "_" + date

				if tempKey not in tempDict:
					tempDict[tempKey] = {}

				cleanVal = "" if (value == "" or value is None or str(value).startswith("#")) else float(value)
				if(indicator == "federal_public_employment" or indicator == "private_employment" or indicator == "public_employment" or indicator == "state_and_local_public_employment" or indicator == "total_employment" or indicator == "state_and_local_public_education_employment" or indicator == "leisure_and_hospitality_employment" or indicator == "manufacturing_employment" or indicator == "retail_trade_employment" ):
					cleanVal *= 1000
				elif(indicator == "state_gdp" or indicator == "accommodation_and_food_services_state_gdp" or indicator == "retail_trade_state_gdp" or indicator == "government_state_gdp" or indicator == "manufacturing_state_gdp" ):
					cleanVal *= 1000000
				tempDict[tempKey][key] = cleanVal

for index, indicator in enumerate(indicators):
	if(indicator != "unemployment_rate"):
		key = str(index) + "c"

		with open(rootPath + "static/data/csv/%s_yoy_percent_change.csv" % indicator, 'r') as changeCsvFile:
			changeReader = csv.reader(changeCsvFile)
			rows = list(changeReader)

		dates = rows[0]
		terminalDates[key] = {"firstDate" : dates[1], "lastDate" : dates[len(dates) - 1]}

		for row in rows[1:]:
			name = row[0]

			for i in range(1, len(row)):
				date = dates[i]
				abbr = nameToFips[name]["abbr"]
				value = row[i]
				tempKey = abbr + "_" + date
				if tempKey not in tempDict:
					tempDict[tempKey] = {}

				cleanVal = "" if (value == "" or value is None or str(value).startswith("#")) else float(value)
				tempDict[tempKey][key] = cleanVal


dataOut = {"data": []}
tempList = []

for tk in tempDict:
	date = tk.split("_")[1]
	abbr = tk.split("_")[0]
	obj = tempDict[tk]
	obj["date"] = date
	obj["abbr"] = abbr
	tempList.append(obj)

dataOut["data"] = sorted(tempList, key=lambda k: datetime.datetime.strptime(k['date'], DATE_FORMAT))

with open(rootPath + 'static/data/cards/cards.json') as inFile:
    cardData = load(inFile)
    dataOut["cards"] = cardData

dataOut["terminalDates"] = terminalDates

now = datetime.datetime.now()

dataOut["lastUpdated"] = now.strftime("%B %d, %Y")


quarterly = ["house_price_index_yoy_percent_change.csv","state_gdp_raw_in_millions.csv","state_gdp_yoy_percent_change.csv","accommodation_and_food_services_state_gdp_raw_in_millions.csv", "accommodation_and_food_services_state_gdp_yoy_percent_change.csv", "retail_trade_state_gdp_raw_in_millions.csv", "retail_trade_state_gdp_yoy_percent_change.csv", "government_state_gdp_raw_in_millions.csv", "government_state_gdp_yoy_percent_change.csv", "manufacturing_state_gdp_raw_in_millions.csv", "manufacturing_state_gdp_yoy_percent_change.csv"]

for filename in quarterly:
	tempfile = NamedTemporaryFile(mode='w', delete=False, newline='')

	with open(rootPath + 'static/data/csv/' + filename, 'r', newline='') as csvFile, tempfile:
		reader = csv.reader(csvFile, delimiter=',', quotechar='"')
		writer = csv.writer(tempfile, delimiter=',', quotechar='"')
		dates = next(reader)
		newDates = []
		for i, r in enumerate(dates):
			if r == "Geography":
				newDates.append(r)
				continue
			tempDate = r.split("-")
			tempYear = tempDate[0]
			tempMonth = int(tempDate[1])
			tempQ = ((tempMonth-1)//3) + 1
			newDates.append(tempYear + " Q" + str(tempQ))

		writer.writerow(newDates)

		for row in reader:
			writer.writerow(row)

	shutil.move(tempfile.name, rootPath + 'static/data/csv/' + filename)



#write a pretty printed json for human readability
with open(rootPath + 'static/data/figures/pretty.json', 'wt') as out:
    res = dump(dataOut, out, sort_keys=True, indent=4, separators=(',', ': '))


#write a one line json for consumption in JS
with open(rootPath + 'static/data/figures/data.json', 'wt') as out:
    res = dump(dataOut, out, sort_keys=True, separators=(',', ':'))
